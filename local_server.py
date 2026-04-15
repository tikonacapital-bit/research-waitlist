import json
import mimetypes
import os
from datetime import datetime, timezone
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib import error, request
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parent
PUBLIC_DIR = BASE_DIR / "public"
ASSETS_DIR = BASE_DIR / "assets"
DATA_DIR = BASE_DIR / "data"
WORKBOOK_PATH = DATA_DIR / "waitlist.xlsx"
HOST = os.environ.get("HOST", "127.0.0.1")
PORT = int(os.environ.get("PORT", 3000))
HEADERS = ["Timestamp", "Email", "Phone", "Source"]
ENV_PATH = BASE_DIR / ".env"


def load_dotenv():
    if not ENV_PATH.exists():
        return

    for raw_line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")

        if key and key not in os.environ:
            os.environ[key] = value


load_dotenv()

STORAGE_MODE = os.getenv("STORAGE_MODE", "excel").strip().lower()
GOOGLE_SHEETS_WEBHOOK_URL = os.getenv("GOOGLE_SHEETS_WEBHOOK_URL", "").strip()


def ensure_workbook():
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    if WORKBOOK_PATH.exists():
        workbook = load_workbook(WORKBOOK_PATH)
        sheet = workbook.active
        return workbook, sheet

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Waitlist"
    sheet.append(HEADERS)
    return workbook, sheet


def autosize_columns(sheet):
    for column_cells in sheet.columns:
        longest = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = max(16, min(longest + 4, 40))


def append_signup(email, phone):
    workbook, sheet = ensure_workbook()
    sheet.append(
        [
            datetime.now(timezone.utc).isoformat(),
            email,
            phone,
            "website",
        ]
    )
    autosize_columns(sheet)
    workbook.save(WORKBOOK_PATH)


def append_signup_to_google_sheets(email, phone):
    if not GOOGLE_SHEETS_WEBHOOK_URL:
        raise ValueError("Google Sheets webhook is not configured.")

    payload = json.dumps(
        {
            "submittedAt": datetime.now(timezone.utc).isoformat(),
            "email": email,
            "phone": phone,
            "source": "website",
        }
    ).encode("utf-8")
    webhook_request = request.Request(
        GOOGLE_SHEETS_WEBHOOK_URL,
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    try:
        with request.urlopen(webhook_request, timeout=10) as response:
            if response.status >= 400:
                raise RuntimeError("Google Sheets rejected the signup.")
    except error.URLError as exc:
        raise RuntimeError("Unable to reach Google Sheets.") from exc


def store_signup(email, phone):
    if STORAGE_MODE == "google":
        append_signup_to_google_sheets(email, phone)
        return

    if STORAGE_MODE == "google+excel":
        append_signup_to_google_sheets(email, phone)
        append_signup(email, phone)
        return

    append_signup(email, phone)


def validate_storage_configuration():
    valid_modes = {"excel", "google", "google+excel"}

    if STORAGE_MODE not in valid_modes:
        raise ValueError(
            f"Invalid STORAGE_MODE '{STORAGE_MODE}'. Use excel, google, or google+excel."
        )

    if STORAGE_MODE in {"google", "google+excel"}:
        if not GOOGLE_SHEETS_WEBHOOK_URL:
            raise ValueError(
                "GOOGLE_SHEETS_WEBHOOK_URL is missing. Paste your Google Apps Script Web App URL."
            )

        if GOOGLE_SHEETS_WEBHOOK_URL == "PASTE_YOUR_WEB_APP_URL_HERE":
            raise ValueError(
                "GOOGLE_SHEETS_WEBHOOK_URL still has the placeholder value. Replace it with the real Web App URL."
            )

        if not GOOGLE_SHEETS_WEBHOOK_URL.startswith("https://script.google.com/"):
            raise ValueError(
                "GOOGLE_SHEETS_WEBHOOK_URL must be your deployed Google Apps Script Web App URL."
            )


def is_valid_email(email):
    if "@" not in email or "." not in email.split("@")[-1]:
        return False
    return " " not in email and len(email) >= 5


def is_valid_phone(phone):
    allowed = set("0123456789+-() ")
    digits_only = "".join(char for char in phone if char.isdigit())
    return 7 <= len(digits_only) <= 12 and all(char in allowed for char in phone)


class WaitlistHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(PUBLIC_DIR), **kwargs)

    def do_GET(self):
        parsed = urlparse(self.path)

        if parsed.path == "/api/download":
            self.handle_download()
            return

        if parsed.path.startswith("/assets/"):
            self.serve_asset(parsed.path.replace("/assets/", "", 1))
            return

        if parsed.path == "/":
            self.path = "/index.html"
        else:
            self.path = parsed.path

        super().do_GET()

    def do_POST(self):
        parsed = urlparse(self.path)

        if parsed.path != "/api/waitlist":
            self.send_error(HTTPStatus.NOT_FOUND, "Not found")
            return

        content_length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(content_length).decode("utf-8")

        try:
            payload = json.loads(body or "{}")
        except json.JSONDecodeError:
            self.send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "message": "Invalid request payload."})
            return

        email = str(payload.get("email", "")).strip()
        phone = str(payload.get("phone", "")).strip()

        if not is_valid_email(email):
            self.send_json(HTTPStatus.BAD_REQUEST, {"ok": False, "message": "Please enter a valid email address."})
            return

        if phone and not is_valid_phone(phone):
            self.send_json(
                HTTPStatus.BAD_REQUEST,
                {
                    "ok": False,
                    "message": "Phone number must be 7 to 12 digits and can include spaces, brackets, plus and hyphen.",
                },
            )
            return

        try:
            store_signup(email, phone)
        except ValueError as exc:
            self.send_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"ok": False, "message": str(exc)})
            return
        except RuntimeError as exc:
            self.send_json(HTTPStatus.BAD_GATEWAY, {"ok": False, "message": str(exc)})
            return

        message = "You're on the list. We'll be in touch soon."
        if STORAGE_MODE == "google":
            message = "Saved to Google Sheet."
        if STORAGE_MODE == "google+excel":
            message = "You're on the list. Your details were saved to Google Sheets."

        self.send_json(HTTPStatus.OK, {"ok": True, "message": message})

    def handle_download(self):
        if not WORKBOOK_PATH.exists():
            self.send_json(HTTPStatus.NOT_FOUND, {"ok": False, "message": "No signups yet."})
            return

        file_bytes = WORKBOOK_PATH.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.send_header("Content-Disposition", 'attachment; filename="waitlist.xlsx"')
        self.send_header("Content-Length", str(len(file_bytes)))
        self.end_headers()
        self.wfile.write(file_bytes)

    def serve_asset(self, relative_path):
        asset_path = (ASSETS_DIR / relative_path).resolve()

        if not str(asset_path).startswith(str(ASSETS_DIR.resolve())) or not asset_path.exists():
            self.send_error(HTTPStatus.NOT_FOUND, "Not found")
            return

        file_bytes = asset_path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", self.guess_type(str(asset_path)))
        self.send_header("Content-Length", str(len(file_bytes)))
        self.end_headers()
        self.wfile.write(file_bytes)

    def send_json(self, status, payload):
        encoded = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(encoded)))
        self.end_headers()
        self.wfile.write(encoded)

    def guess_type(self, path):
        mime_type, _ = mimetypes.guess_type(path)
        return mime_type or "application/octet-stream"


if __name__ == "__main__":
    validate_storage_configuration()
    server = ThreadingHTTPServer((HOST, PORT), WaitlistHandler)
    print(f"Tikona waitlist is running at http://{HOST}:{PORT}")
    print(f"Storage mode: {STORAGE_MODE}")
    if STORAGE_MODE in {"google", "google+excel"}:
        print(f"Google Sheets webhook configured: {GOOGLE_SHEETS_WEBHOOK_URL}")
    server.serve_forever()
